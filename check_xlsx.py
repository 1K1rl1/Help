#!/usr/bin/env python3
import openpyxl
from pathlib import Path

xlsx_path = Path("test_output.xlsx")

if xlsx_path.exists():
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        
        print(f"File: {xlsx_path}")
        print(f"Rows: {ws.max_row}")
        print(f"Columns: {ws.max_column}")
        print()
        print("First 5 rows:")
        print("-" * 200)
        
        for row_idx in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data.append(str(cell.value)[:20] if cell.value else "")
            print(f"Row {row_idx}: {' | '.join(row_data)}")
        
        wb.close()
    except Exception as e:
        print(f"Error reading file: {e}")
else:
    print(f"File not found: {xlsx_path}")
