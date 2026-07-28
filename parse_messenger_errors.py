import os
import re
import time
import logging
import sqlite3
import datetime
import io
import json
from typing import Any, Dict, List, Optional

import requests
from flask import Flask, request, jsonify
from flask import send_file
from dotenv import load_dotenv

print('DEBUG parse_messenger_errors.py loaded from', __file__)
print('DEBUG cwd', os.getcwd())

load_dotenv()

LOG = logging.getLogger(__name__)
OFFLINE_ONLY = os.getenv("OFFLINE_ONLY", "true").lower() in ("1", "true", "yes")

VALID_OPERATIONS = {
    "Размещение": "Потеря при размещении",
    "Размещение_товара": "Размещение_товара",
    "Отпуск": "Отпуск",
    "Приём": "Приём",
    "Прием_товара": "Прием_товара",
}

PROCESS_ALIASES = {
    "прием": "Прием_товара",
    "приемка": "Прием_товара",
    "приём": "Прием_товара",
    "прием товара": "Прием_товара",
    "прием_товара": "Прием_товара",
    "размещение": "Размещение_товара",
    "размес": "Размещение_товара",
    "размещение товара": "Размещение_товара",
    "размещение_товара": "Размещение_товара",
    "разм": "Размещение_товара",
}

PROCESS_TO_ERROR_NAME = {
    "Прием_товара": "Нарушение технологии приема товара",
    "Размещение_товара": "Нарушение технологии размещения",
}

FIELD_PATTERNS = {
    "operation": re.compile(r"^(Операция|Название процесса):\s*(.+)$", re.IGNORECASE),
    "login": re.compile(r"^(Сотрудник\(логин\)|Логин оператора склада):\s*(.+)$", re.IGNORECASE),
    "object_id": re.compile(r"^(Идентификатор объекта|Идентификатор объекта нарушения):\s*(.+)$", re.IGNORECASE),
    "place": re.compile(r"^Место операции:\s*(.+)$", re.IGNORECASE),
    "comment": re.compile(r"^Комментарий:\s*(.*)$", re.IGNORECASE),
    "manager_name": re.compile(r"^(ФИО Начальника Участка|ФИО кто выставил):\s*(.*)$", re.IGNORECASE),
}

EXPECTED_HEADERS = [
    "Объект",
    "Дата фиксации",
    "Название процесса",
    "Название ошибки",
    "Логин оператора склада",
    "Метод начисления ошибки",
    "Идентификатор объекта нарушения",
    "Место операции",
    "Комментарий",
    "ФИО кто выставил",
]

LOGIN_PATTERN = re.compile(r"^[A-Za-z0-9]{3,20}$")
OBJECT_ID_PATTERN = re.compile(r"^[A-Za-z0-9_./=+&()\\-]+$")

GRAPH_SCOPE = "https://graph.microsoft.com/.default"
DEFAULT_DEDUPE_DB = os.getenv("DEDUPE_DB_PATH", "dedupe.db")

_graph_token_cache = {"access_token": None, "expires_at": datetime.datetime.min}


class ValidationError(Exception):
    def __init__(self, errors: List[str]):
        self.errors = errors
        super().__init__("; ".join(errors))


def configure_logging() -> None:
    handler = logging.StreamHandler()
    handler.setFormatter(
        logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S")
    )
    LOG.addHandler(handler)
    LOG.setLevel(logging.INFO)


def retry_call(fn, *args, retries: int = 3, initial_delay: float = 1.0, backoff: float = 2.0, **kwargs):
    delay = initial_delay
    for attempt in range(1, retries + 1):
        try:
            return fn(*args, **kwargs)
        except Exception as exc:
            if attempt == retries:
                LOG.error("Retry exhausted for %s: %s", getattr(fn, "__name__", str(fn)), exc)
                raise
            LOG.warning("Retry %s/%s for %s after error: %s", attempt, retries, getattr(fn, "__name__", str(fn)), exc)
            time.sleep(delay)
            delay *= backoff


def normalize_process_name(value: str) -> str:
    if not value:
        return ""
    normalized = re.sub(r"\s+", " ", value.strip().lower())
    normalized = normalized.replace('ё', 'е')
    return PROCESS_ALIASES.get(normalized, value.strip())


def parse_object_ids(value: str) -> List[str]:
    if not value:
        return []
    # Split on commas, semicolons, newlines and also on whitespace-separated IDs
    candidates = re.split(r"[,;\n]+", value)
    # additionally split space-separated entries when no commas/semicolons are present
    if len(candidates) == 1:
        space_split = re.split(r"\s+", candidates[0])
        if len(space_split) > 1:
            candidates = space_split
    ids = []
    for item in candidates:
        clean = item.strip()
        if not clean:
            continue
        # Keep ii.../bx... prefixes — do not remove them
        clean = re.sub(r"\s+", " ", clean)
        clean = clean.strip(";, ")
        if clean:
            ids.append(clean)
    return ids


def parse_message(text: str) -> Dict[str, str]:
    values: Dict[str, str] = {}
    normalized_text = (
        text
        .replace('\r\n', '\n')
        .replace('\r', '\n')
        .replace('\ufeff', '')
        .replace('：', ':')
        .replace('\u00a0', ' ')
        .replace('\u200b', '')
    )
    LOG.info("parse_message raw text (%d chars): %r", len(text), text)
    LOG.info("parse_message normalized text (%d chars): %r", len(normalized_text), normalized_text)
    alias_to_field = {
        'операция': 'operation',
        'название процесса': 'operation',
        'сотрудник(логин)': 'login',
        'логин оператора склада': 'login',
        'логин': 'login',
        'идентификатор объекта': 'object_id',
        'идентификатор объекта нарушения': 'object_id',
        'идентификатор': 'object_id',
        'место операции': 'place',
        'комментарий': 'comment',
        'фио начальника участка': 'manager_name',
        'фио кто выставил': 'manager_name',
        'фио': 'manager_name',
    }

    current_field: Optional[str] = None
    multiline_fields = {"object_id", "comment", "manager_name", "place"}

    for line in normalized_text.splitlines():
        line = line.strip()
        if not line:
            current_field = None
            continue
        line = re.sub(r"\s+", " ", line)
        LOG.info("parse_message processing line: %r", line)

        # If we're currently collecting a multiline field, and this line doesn't look like a new label,
        # treat it as continuation of the current field value.
        if current_field in multiline_fields:
            is_label = False
            # check for colon-style label
            if ':' in line:
                label_candidate = line.split(':', 1)[0]
                normalized_label = re.sub(r"\s+", " ", label_candidate.strip().lower()).replace('ё', 'е')
                if normalized_label in alias_to_field:
                    is_label = True
            else:
                lowline = line.lower()
                for alias in alias_to_field:
                    na = re.sub(r"\s+", " ", alias.lower()).replace('ё', 'е')
                    if lowline == na or lowline.startswith(na + ' '):
                        is_label = True
                        break
            if not is_label:
                existing = values.get(current_field, "")
                values[current_field] = f"{existing}\n{line}" if existing else line
                LOG.info("parse_message appended (continuation) to %s => %r", current_field, values[current_field])
                continue

        if ':' in line:
            label, value = line.split(':', 1)
            normalized_label = re.sub(r"\s+", " ", label.strip().lower()).replace('ё', 'е')
            LOG.info("parse_message fallback label=%r normalized=%r", label.strip(), normalized_label)
            matched_field = None
            if normalized_label in alias_to_field:
                matched_field = alias_to_field[normalized_label]
            else:
                for alias, field in alias_to_field.items():
                    normalized_alias = re.sub(r"\s+", " ", alias.lower()).replace('ё', 'е')
                    if normalized_alias == normalized_label or normalized_alias in normalized_label or normalized_label in normalized_alias:
                        matched_field = field
                        break

            if matched_field:
                values[matched_field] = value.strip()
                current_field = matched_field if matched_field in multiline_fields else None
                LOG.info("parse_message fallback matched %s => %r", matched_field, values[matched_field])
                continue
            current_field = None
        else:
            # If the line contains only the label (no colon, no value), treat the next lines as the field value
            label_only_matched = False
            for alias, field in alias_to_field.items():
                try:
                    label_only_pattern = re.compile(rf"^{re.escape(alias)}$", re.IGNORECASE)
                except re.error:
                    continue
                if label_only_pattern.match(line):
                    # initialize empty value and set current_field to collect following lines
                    values.setdefault(field, "")
                    current_field = field if field in multiline_fields else None
                    LOG.info("parse_message detected label-only %s, switching current_field=%s", alias, current_field)
                    label_only_matched = True
                    break
            if label_only_matched:
                continue

            # Try matching patterns like "Название процесса размещение" (no colon)
            matched_field = None
            # Prefer longest aliases first to avoid partial matches
            for alias, field in sorted(alias_to_field.items(), key=lambda x: -len(x[0])):
                try:
                    pattern = re.compile(rf"^{re.escape(alias)}\s+(.+)", re.IGNORECASE)
                except re.error:
                    continue
                m = pattern.match(line)
                if m:
                    values[field] = m.group(1).strip()
                    current_field = field if field in multiline_fields else None
                    LOG.info("parse_message matched no-colon %s => %r", field, values[field])
                    matched_field = field
                    break
            if matched_field:
                continue
            current_field = None

        if current_field in multiline_fields:
            existing = values.get(current_field, "")
            values[current_field] = f"{existing}\n{line}" if existing else line
            LOG.info("parse_message appended line to %s => %r", current_field, values[current_field])
            continue

        for key, pattern in FIELD_PATTERNS.items():
            match = pattern.match(line) or pattern.search(line)
            if match:
                values[key] = match.group(match.lastindex).strip()
                current_field = key if key in multiline_fields else None
                LOG.info("parse_message matched %s => %r", key, values[key])
                break

    if not values.get('operation') or not values.get('login') or not values.get('object_id'):
        for field, aliases in {
            'operation': ['операция', 'название процесса'],
            'login': ['сотрудник(логин)', 'логин оператора склада'],
            'object_id': ['идентификатор объекта', 'идентификатор объекта нарушения'],
        }.items():
            if values.get(field):
                continue
            for alias in aliases:
                regex = re.compile(rf"{re.escape(alias)}\s*[:]\s*(.+)", re.IGNORECASE)
                match = regex.search(normalized_text)
                if match:
                    values[field] = match.group(1).strip()
                    LOG.info("parse_message fallback whole text matched %s => %r", field, values[field])
                    break

    # fallback: search whole text for missing values by alias if line parsing failed
    for field, aliases in {
        'operation': ['операция', 'название процесса'],
        'login': ['сотрудник(логин)', 'логин оператора склада'],
        'object_id': ['идентификатор объекта', 'идентификатор объекта нарушения'],
    }.items():
        if values.get(field):
            continue
        for alias in aliases:
            regex = re.compile(rf"{re.escape(alias)}\s*[:]\s*(.+)", re.IGNORECASE)
            match = regex.search(normalized_text)
            if match:
                values[field] = match.group(1).strip()
                LOG.info("parse_message fallback whole text matched %s => %r", field, values[field])
                break

    if values.get('object_id'):
        parsed_ids = parse_object_ids(values['object_id'])
        if parsed_ids:
            values['object_ids'] = parsed_ids
            values['object_id'] = parsed_ids[0]

    if values.get('operation'):
        normalized_operation = normalize_process_name(values['operation'])
        values['operation'] = normalized_operation

    if values.get('place'):
        values['place'] = 'ОПиР'

    if values.get('comment') and not values.get('error_name'):
        values['error_name'] = PROCESS_TO_ERROR_NAME.get(values.get('operation', ''), '')

    if values.get('operation') and not values.get('error_name'):
        values['error_name'] = PROCESS_TO_ERROR_NAME.get(values.get('operation', ''), '')

    LOG.info("parse_message extracted values: %s", values)

    errors: List[str] = []

    operation = (values.get("operation") or "").strip()
    login = (values.get("login") or "").strip()
    comment = (values.get("comment") or "").strip()
    manager_name = (values.get("manager_name") or "").strip()
    object_id = (values.get("object_id") or "").strip()

    if not operation:
        errors.append("Отсутствует обязательное поле: Название процесса")
    elif operation.lower() in {"input", ""}:
        errors.append("Название процесса не может быть пустым")
    else:
        # Accept various aliases and normalized forms for operation
        normalized_check = normalize_process_name(operation) or operation
        allowed_ops = set(VALID_OPERATIONS.keys()) | set(PROCESS_ALIASES.values())
        if normalized_check not in allowed_ops:
            errors.append("Недопустимое значение для Название процесса. Используйте: размещение или приемка")

    if not login:
        errors.append("Отсутствует обязательное поле: Логин оператора склада")
    elif not LOGIN_PATTERN.match(login):
        errors.append("Неверный формат логина. Допускаются латинские буквы и цифры, 3-20 символов.")

    if not comment:
        errors.append("Отсутствует обязательное поле: Комментарий")

    if not manager_name:
        errors.append("Отсутствует обязательное поле: ФИО кто выставил")

    parsed_ids = values.get("object_ids") or parse_object_ids(object_id)
    if parsed_ids:
        values["object_ids"] = parsed_ids
        values["object_id"] = parsed_ids[0]
        for candidate in parsed_ids:
            if not candidate or not OBJECT_ID_PATTERN.fullmatch(candidate):
                errors.append("Неверный формат идентификатора объекта. Допускаются латинские буквы, цифры и подчёркивание.")
                break
    else:
        errors.append("Отсутствует обязательное поле: Идентификатор объекта нарушения")

    if errors:
        raise ValidationError(errors)

    return values


def determine_operation(values: Dict[str, str]) -> str:
    """Определяет ключ операции (`Размещение`, `Приём`, `Отпуск`) по полю operation и heuristics по комментарию."""
    comment = values.get('comment', '') or ''
    operation_raw = (values.get('operation') or '').strip()

    # heuristics from comment
    if re.search(r"Принят|Приемка|Положил не в ту тару", comment, re.IGNORECASE):
        return 'Приём'
    if re.search(r"Потеря|По маршруту", comment, re.IGNORECASE):
        return 'Размещение'

    # if user provided a known operation, use it
    if operation_raw and operation_raw in VALID_OPERATIONS:
        return operation_raw

    # fallback: if raw equals common English marker 'input' or similar, leave as-is
    return operation_raw


def build_table_row(values: Dict[str, Any], object_id: str, timestamp: Optional[datetime.datetime] = None) -> List[str]:
    # Формируем строку в порядке, указанном пользователем:
    # [sheet_owner, Дата и время совершения ошибки, Название процесса, Название ошибки, Логин оператора склада,
    #  Метод начисления ошибки, Идентификатор объекта нарушения, Место операции, Комментарий, ФИО кто выставил]
    sheet_owner = os.getenv("SHEET_OWNER", "Екатеринбург_РФЦ_НОВЫЙ")
    comment = values.get("comment", "")
    manager_name = values.get("manager_name", "")
    login = values.get("login", "")

    raw_op = (values.get("operation") or "").strip()
    if not raw_op or raw_op.lower() == 'input':
        process_name = determine_operation(values)
    else:
        process_name = raw_op

    error_name = values.get("error_name") or PROCESS_TO_ERROR_NAME.get(process_name, "")
    timestamp = timestamp or datetime.datetime.now()
    timestamp_str = timestamp.strftime("%Y-%m-%d %H:%M:%S")

    # order: Объект, Дата фиксации, Название процесса, Название ошибки, Логин оператора склада,
    # Метод начисления ошибки, Идентификатор объекта нарушения, Место операции, Комментарий, ФИО кто выставил
    row = [
        sheet_owner,
        timestamp_str,
        process_name,
        error_name,
        login,
        "Ручной",
        object_id,
        values.get("place", "ОПиР"),
        comment,
        manager_name,
    ]
    return row


def build_table_rows(values: Dict[str, Any]) -> List[List[str]]:
    object_ids = values.get("object_ids") or [values.get("object_id", "")]
    base_time = datetime.datetime.now()
    rows: List[List[str]] = []
    for idx, object_id in enumerate(object_ids):
        rows.append(build_table_row(values, object_id, base_time + datetime.timedelta(seconds=idx)))
    return rows


def format_table_row(row: List[str]) -> str:
    """Форматирует строку таблицы в красивый текстовый вывод с правильным порядком полей."""
    headers = [
        "Объект",
        "Дата фиксации",
        "Название процесса",
        "Название ошибки",
        "Логин оператора склада",
        "Метод начисления ошибки",
        "Идентификатор объекта нарушения",
        "Место операции",
        "Комментарий",
        "ФИО кто выставил",
    ]
    
    formatted_lines = []
    for header, value in zip(headers, row):
        formatted_lines.append(f"{header} → {value}")
    
    return "\n".join(formatted_lines)


def ensure_dedupe_db(db_path: str) -> None:
    conn = sqlite3.connect(db_path)
    try:
        cursor = conn.cursor()
        cursor.execute(
            "CREATE TABLE IF NOT EXISTS submissions (id INTEGER PRIMARY KEY AUTOINCREMENT, object_id TEXT NOT NULL, created_at TIMESTAMP NOT NULL, raw_message TEXT)"
        )
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_object_time ON submissions(object_id, created_at)")
        conn.commit()
    finally:
        conn.close()

def ensure_rows_table(db_path: str) -> None:
    conn = sqlite3.connect(db_path)
    try:
        cursor = conn.cursor()
        cursor.execute(
            "CREATE TABLE IF NOT EXISTS rows (id INTEGER PRIMARY KEY AUTOINCREMENT, object_id TEXT NOT NULL, created_at TIMESTAMP NOT NULL, row_json TEXT)"
        )
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_rows_object ON rows(object_id)")
        conn.commit()
    finally:
        conn.close()


def is_duplicate(db_path: str, object_id: str, window_hours: int = 24) -> bool:
    conn = sqlite3.connect(db_path)
    try:
        cursor = conn.cursor()
        threshold = datetime.datetime.now() - datetime.timedelta(hours=window_hours)
        cursor.execute("SELECT COUNT(1) FROM submissions WHERE object_id = ? AND created_at >= ?", (object_id, threshold.isoformat()))
        count = cursor.fetchone()[0]
        return count > 0
    finally:
        conn.close()


def record_submission(db_path: str, object_id: str, raw_message: str) -> None:
    conn = sqlite3.connect(db_path)
    try:
        cursor = conn.cursor()
        cursor.execute("INSERT INTO submissions (object_id, created_at, raw_message) VALUES (?, ?, ?)", (object_id, datetime.datetime.now().isoformat(), raw_message))
        conn.commit()
    finally:
        conn.close()

def write_row_to_db(db_path: str, row: List[str]) -> None:
    ensure_rows_table(db_path)
    conn = sqlite3.connect(db_path)
    try:
        cursor = conn.cursor()
        object_id = row[6] if len(row) > 6 else ''
        cursor.execute("INSERT INTO rows (object_id, created_at, row_json) VALUES (?, ?, ?)", (object_id, datetime.datetime.now().isoformat(), json.dumps(row, ensure_ascii=False)))
        conn.commit()
    finally:
        conn.close()

def read_existing_db_ids(db_path: str) -> List[str]:
    if not os.path.exists(db_path):
        return []
    conn = sqlite3.connect(db_path)
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT object_id FROM rows")
        ids = [str(r[0]) for r in cursor.fetchall() if r and r[0]]
        return ids
    finally:
        conn.close()

def export_rows_to_xlsx(db_path: str) -> io.BytesIO:
    try:
        from openpyxl import Workbook
    except ImportError:
        raise RuntimeError('openpyxl is required to export XLSX')
    conn = sqlite3.connect(db_path)
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT row_json FROM rows ORDER BY created_at ASC")
        rows = [json.loads(r[0]) for r in cursor.fetchall() if r and r[0]]
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.append(EXPECTED_HEADERS)
    for row in rows:
        try:
            ws.append([str(cell) for cell in row])
        except Exception:
            ws.append(row)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def get_graph_access_token(tenant_id: str, client_id: str, client_secret: str) -> str:
    now = datetime.datetime.now()
    token = _graph_token_cache.get("access_token")
    if token and _graph_token_cache.get("expires_at") and _graph_token_cache["expires_at"] > now + datetime.timedelta(seconds=30):
        return token

    token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    resp = requests.post(
        token_url,
        data={
            "grant_type": "client_credentials",
            "client_id": client_id,
            "client_secret": client_secret,
            "scope": GRAPH_SCOPE,
        },
        timeout=10,
    )
    if not resp.ok:
        LOG.error("Graph token request failed %s: %s", resp.status_code, resp.text)
    resp.raise_for_status()
    data = resp.json()
    access_token = data["access_token"]
    expires_in = int(data.get("expires_in", 3600))
    _graph_token_cache["access_token"] = access_token
    _graph_token_cache["expires_at"] = now + datetime.timedelta(seconds=expires_in)
    return access_token


def append_row_to_excel(drive_id: str, item_id: str, table_name: str, row: List[str], tenant_id: str, client_id: str, client_secret: str) -> None:
    access_token = get_graph_access_token(tenant_id, client_id, client_secret)
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/workbook/tables/{table_name}/rows/add"
    headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}
    body = {"values": [row]}
    resp = requests.post(url, headers=headers, json=body, timeout=20)
    resp.raise_for_status()
    LOG.info("Добавлена строка в Excel Online: %s", row)


def upload_local_file_to_drive(drive_id: str, item_id: str, file_path: str, tenant_id: str, client_id: str, client_secret: str) -> str:
    access_token = get_graph_access_token(tenant_id, client_id, client_secret)
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content"
    headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/octet-stream"}
    with open(file_path, "rb") as f:
        resp = requests.put(url, headers=headers, data=f, timeout=60)
    if not resp.ok:
        LOG.error("Graph upload content failed %s: %s", resp.status_code, resp.text)
    resp.raise_for_status()
    data = resp.json()
    web_url = data.get("webUrl")
    if web_url:
        return web_url
    return get_drive_item_web_url(drive_id, item_id, tenant_id, client_id, client_secret)


def get_drive_item_web_url(drive_id: str, item_id: str, tenant_id: str, client_id: str, client_secret: str) -> str:
    access_token = get_graph_access_token(tenant_id, client_id, client_secret)
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}?select=webUrl"
    headers = {"Authorization": f"Bearer {access_token}"}
    resp = requests.get(url, headers=headers, timeout=20)
    if not resp.ok:
        LOG.error("Graph get item webUrl failed %s: %s", resp.status_code, resp.text)
    resp.raise_for_status()
    data = resp.json()
    web_url = data.get("webUrl")
    if web_url:
        return web_url
    return create_share_link(drive_id, item_id, tenant_id, client_id, client_secret)


def create_share_link(drive_id: str, item_id: str, tenant_id: str, client_id: str, client_secret: str) -> str:
    access_token = get_graph_access_token(tenant_id, client_id, client_secret)
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/createLink"
    headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}
    body = {"type": "view", "scope": "organization"}
    resp = requests.post(url, headers=headers, json=body, timeout=20)
    if not resp.ok:
        LOG.error("Graph create share link failed %s: %s", resp.status_code, resp.text)
    resp.raise_for_status()
    data = resp.json()
    link = data.get("link", {}).get("webUrl")
    if not link:
        raise RuntimeError("Не удалось получить ссылку для общего доступа к файлу")
    return link


def read_existing_csv_ids(file_path: str) -> List[str]:
    import csv
    if not os.path.exists(file_path):
        return []
    ids = []
    with open(file_path, "r", newline="", encoding=os.getenv("TEST_CSV_ENCODING", "utf-8-sig")) as f:
        reader = csv.reader(f)
        headers = next(reader, [])
        if headers[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
            return []
        for row in reader:
            row_values = [str(cell).strip() for cell in row[: len(EXPECTED_HEADERS)]]
            if is_empty_excel_row(row_values):
                continue
            ids.append(row_values[6])
    return [obj_id for obj_id in ids if obj_id]


def write_row_to_csv(file_path: str, row: List[str]) -> None:
    import csv
    exists = os.path.exists(file_path)
    encoding = os.getenv("TEST_CSV_ENCODING", "utf-8-sig")
    # Use utf-8-sig by default so Excel recognizes UTF-8 with BOM on Windows
    with open(file_path, "a", newline="", encoding=encoding) as f:
        writer = csv.writer(f)
        if not exists:
            writer.writerow(EXPECTED_HEADERS)
        writer.writerow(row)
    LOG.info("Записана тестовая строка в %s (encoding=%s): %s", file_path, encoding, row)


def normalize_excel_cell(value: Optional[str]) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_empty_excel_row(row_values: List[str]) -> bool:
    return all(not cell.strip() for cell in row_values)


def read_existing_excel_ids(file_path: str) -> List[str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    if not os.path.exists(file_path):
        return []

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    ids = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            headers = [normalize_excel_cell(cell) for cell in row]
            if headers[: len(EXPECTED_HEADERS)] == EXPECTED_HEADERS:
                header_found = True
            continue
        row_values = [normalize_excel_cell(cell) for cell in row[: len(EXPECTED_HEADERS)]]
        if is_empty_excel_row(row_values):
            continue
        ids.append(row_values[6])
    wb.close()
    return [obj_id for obj_id in ids if obj_id]


def clean_empty_rows(ws) -> None:
    max_col = len(EXPECTED_HEADERS)
    for row_idx in range(ws.max_row, 1, -1):
        row_cells = [cell.value for cell in ws[row_idx][:max_col]]
        row_values = [normalize_excel_cell(value) for value in row_cells]
        if is_empty_excel_row(row_values):
            ws.delete_rows(row_idx, 1)


def write_row_to_xlsx(file_path: str, row: List[str]) -> None:
    try:
        from openpyxl import Workbook, load_workbook
    except Exception:
        raise

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(EXPECTED_HEADERS)

    if ws.max_row == 0:
        ws.append(EXPECTED_HEADERS)
    else:
        expected_col_count = len(EXPECTED_HEADERS)
        if ws.max_column < expected_col_count:
            for idx in range(ws.max_column + 1, expected_col_count + 1):
                ws.cell(row=1, column=idx, value=EXPECTED_HEADERS[idx - 1])

    clean_empty_rows(ws)
    ws.append(row)
    wb.save(file_path)
    LOG.info("Записана тестовая строка в XLSX %s: %s", file_path, row)


def notify_max(notify_url: str, payload: Dict) -> None:
    try:
        retry_call(lambda: requests.post(notify_url, json=payload, timeout=10))
        LOG.info("Отправлено уведомление в MAX: %s", payload)
    except Exception as exc:
        LOG.exception("Не удалось отправить уведомление в MAX: %s", exc)


app = Flask(__name__)

@app.route("/upload_offline", methods=["POST"])
def upload_offline():
    data = request.get_json(silent=True) or {}
    drive_id = os.getenv("EXCEL_DRIVE_ID")
    item_id = os.getenv("EXCEL_ITEM_ID")
    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    file_path = os.getenv("TEST_OUTPUT_XLSX", "test_output.xlsx")

    if not drive_id or not item_id or not tenant_id or not client_id or not client_secret:
        return jsonify({"error": "Excel/Azure configuration не настроена"}), 500
    if not os.path.exists(file_path):
        return jsonify({"error": f"Файл не найден: {file_path}"}), 404

    try:
        link = upload_local_file_to_drive(drive_id, item_id, file_path, tenant_id, client_id, client_secret)
        return jsonify({"status": "ok", "link": link}), 200
    except Exception as exc:
        LOG.exception("Failed to upload offline file to cloud: %s", exc)
        return jsonify({"error": str(exc), "hint": "Проверьте Azure/Excel config и токены Graph API"}), 500


def build_reply_text(status: str, details: Optional[List[str]] = None, object_id: str = "") -> str:
    if status == "ok":
        return "✅ Запись обработана"
    if status == "duplicate":
        return f"⚠️ Дубликат для объекта {object_id}" if object_id else "⚠️ Дубликат"
    if status == "validation_error":
        if details:
            return "❌ Ошибка ввода: " + "; ".join(details)
        return "❌ Ошибка ввода"
    if details:
        return "❌ Ошибка: " + "; ".join(details)
    return "❌ Ошибка"


def send_bot_reply(notify_url: str, chat_id: Optional[str], text: str, extra_payload: Optional[Dict] = None) -> None:
    if not notify_url or not chat_id:
        return
    payload: Dict = {"chat_id": chat_id, "chat": {"id": chat_id}, "text": text}
    if extra_payload:
        payload.update(extra_payload)
    notify_max(notify_url, payload)


# Flask receiver


@app.route("/incoming", methods=["POST"])
def incoming():
    data = request.get_json(silent=True)
    with open("incoming_debug.log", "a", encoding="utf-8") as debug_file:
        debug_file.write("--- INCOMING START ---\n")
        debug_file.write(f"data={data!r}\n")
    if not data:
        return jsonify({"error": "no json body"}), 400

    # message shape forwarded from `max_bot.js`
    text = data.get("text")
    if not isinstance(text, str):
        text = ""
    text = text.strip()
    with open("incoming_debug.log", "a", encoding="utf-8") as debug_file:
        debug_file.write(f"top_level_text={text!r}\n")
    if not text:
        raw_text = data.get("raw", {}).get("body", {}).get("text") or data.get("raw", {}).get("text")
        if isinstance(raw_text, str):
            text = raw_text.strip()
        else:
            text = ""
        if text:
            with open("incoming_debug.log", "a", encoding="utf-8") as debug_file:
                debug_file.write(f"fallback_raw_text={text!r}\n")

    chat = data.get("chat", {}) or {}
    raw = data.get("raw") or {}
    recipient = raw.get("recipient") or {}
    message = {
        "chat": {"id": chat.get("id"), "type": chat.get("type")},
        "text": text,
        "message_id": data.get("message_id"),
    }
    chat_id = message.get("chat", {}).get("id")
    if not chat_id:
        chat_id = recipient.get("chat_id") or data.get("chat_id") or data.get("recipient") or data.get("chat", {}).get("chat_id")
    chat_type = message.get("chat", {}).get("type") or ""
    allowed_chat_ids = {item.strip() for item in os.getenv("ALLOWED_CHAT_IDS", "").split(",") if item.strip()}
    allowed_chat_id = os.getenv("ALLOWED_CHAT_ID")
    if allowed_chat_id:
        allowed_chat_ids.add(allowed_chat_id.strip())

    if allowed_chat_ids and str(chat_id) not in allowed_chat_ids:
        LOG.info("Ignoring chat id %s not listed in ALLOWED_CHAT_IDS/ALLOWED_CHAT_ID", chat_id)
        return jsonify({"status": "ignored", "reason": "chat not allowed"}), 200

    if isinstance(chat_type, str) and chat_type.lower() in ("private", "personal", "direct", "user", "dm", "im"):
        LOG.info("Ignoring private chat type %s for chat_id %s", chat_type, chat_id)
        return jsonify({"status": "ignored", "reason": "private chat"}), 200

    with open("incoming_debug.log", "a", encoding="utf-8") as debug_file:
        debug_file.write(f"final_message_text={text!r}\n")

    drive_id = os.getenv("EXCEL_DRIVE_ID")
    item_id = os.getenv("EXCEL_ITEM_ID")
    table_name = os.getenv("EXCEL_TABLE_NAME", "Table1")
    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    notify_chat_id = os.getenv("NOTIFICATION_CHAT_ID")
    notify_url = os.getenv("NOTIFY_URL")
    db_path = os.getenv("DEDUPE_DB_PATH", DEFAULT_DEDUPE_DB)
    if chat_id is None:
        chat_id = data.get("chat_id") or data.get("recipient") or data.get("chat", {}).get("chat_id")
    if chat_id is None and notify_chat_id:
        chat_id = notify_chat_id
    if isinstance(chat_id, str):
        chat_id = chat_id.strip() or None

    # Do not abort early if Azure/Excel is not configured — fallback to CSV for testing
    if not drive_id or not item_id or not tenant_id or not client_id or not client_secret:
        LOG.warning("Excel/Azure configuration not found — using CSV fallback for testing")

    ensure_dedupe_db(db_path)

    try:
        fields = parse_message(text)
        LOG.info("Parsed message fields: %s", fields)
    except ValidationError as exc:
        LOG.error("Validation error: %s", exc.errors)
        send_bot_reply(notify_url, chat_id, build_reply_text("validation_error", exc.errors), {"reply_to_message_id": message.get("message_id")})
        return jsonify({"status": "validation_error", "errors": exc.errors}), 200

    object_ids = fields.get("object_ids") or [fields.get("object_id")]
    object_ids = [str(obj_id).strip() for obj_id in object_ids if obj_id]
    if not object_ids:
        send_bot_reply(notify_url, chat_id, build_reply_text("validation_error", ["Нет идентификатора объекта."]), {"reply_to_message_id": message.get("message_id")})
        return jsonify({"status": "validation_error", "errors": ["Нет идентификатора объекта."]}), 200

    base_time = datetime.datetime.now()
    rows_to_write: List[Dict[str, object]] = []
    duplicate_ids: List[str] = []
    for idx, oid in enumerate(object_ids):
        if is_duplicate(db_path, oid):
            LOG.info("Duplicate within 24h detected for object_id=%s. No row will be written.", oid)
            duplicate_ids.append(oid)
            continue
        rows_to_write.append({
            "object_id": oid,
            "row": build_table_row(fields, oid, base_time + datetime.timedelta(seconds=idx)),
        })

    if not rows_to_write and duplicate_ids:
        duplicate_text = ", ".join(duplicate_ids)
        send_bot_reply(notify_url, chat_id, build_reply_text("duplicate", object_id=duplicate_text), {"reply_to_message_id": message.get("message_id")})
        return jsonify({"status": "duplicate", "object_ids": duplicate_ids}), 200

    LOG.info("Built output rows: %s", [r["row"] for r in rows_to_write])
    written_ids: List[str] = []

    # If Azure/Excel not configured, write to local CSV/XLSX for testing
    if not drive_id or not item_id or not tenant_id or not client_id or not client_secret:
        # Use SQLite rows table for offline storage
        db_rows_path = db_path
        out_format = os.getenv("TEST_OUTPUT_FORMAT", "xlsx").lower()
        existing_ids = read_existing_db_ids(db_rows_path)
        for oid in object_ids:
            if oid in existing_ids and oid not in duplicate_ids:
                duplicate_ids.append(oid)
        rows_to_write = [item for item in rows_to_write if item["object_id"] not in duplicate_ids]
        try:
            for item in rows_to_write:
                row = item["row"]
                write_row_to_db(db_rows_path, row)
                LOG.info("Wrote row to DB: %s", row)
                record_submission(db_path, item["object_id"], text)
                written_ids.append(str(item["object_id"]))

            status_text = f"✅ Запись сохранена. Добавлено {len(written_ids)} строк(а)."
            if duplicate_ids:
                status_text += f" ⚠️ Пропущено дубликатов: {', '.join(duplicate_ids)}"
            if notify_url:
                send_bot_reply(notify_url, chat_id, status_text, {"status": "ok", "object_ids": written_ids, "reply_to_message_id": message.get("message_id")})
        except Exception as exc:
            LOG.exception("Failed to write test output to DB: %s", exc)
            if notify_url:
                send_bot_reply(notify_url, chat_id, build_reply_text("error", [str(exc)]), {"reply_to_message_id": message.get("message_id")})
            return jsonify({"error": str(exc)}), 500
    else:
        try:
            for item in rows_to_write:
                append_row_to_excel(drive_id, item_id, table_name, item["row"], tenant_id, client_id, client_secret)
                record_submission(db_path, item["object_id"], text)
                written_ids.append(str(item["object_id"]))
            status_text = f"✅ Запись сохранена. Добавлено {len(written_ids)} строк(а)."
            if duplicate_ids:
                status_text += f" ⚠️ Пропущено дубликатов: {', '.join(duplicate_ids)}"
            if notify_url:
                send_bot_reply(notify_url, chat_id, status_text, {"status": "ok", "object_ids": written_ids, "reply_to_message_id": message.get("message_id")})
        except Exception as exc:
            LOG.exception("Failed to append row to Excel: %s", exc)
            if notify_url:
                send_bot_reply(notify_url, chat_id, build_reply_text("error", [str(exc)]), {"reply_to_message_id": message.get("message_id")})
            return jsonify({"error": str(exc)}), 500

    return jsonify({"status": "ok"}), 200


@app.route('/export_offline', methods=['GET'])
def export_offline():
    db_path = os.getenv("DEDUPE_DB_PATH", DEFAULT_DEDUPE_DB)
    try:
        bio = export_rows_to_xlsx(db_path)
        return send_file(bio, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='export.xlsx')
    except Exception as exc:
        LOG.exception('Failed to export rows: %s', exc)
        return jsonify({'error': str(exc)}), 500


if __name__ == "__main__":
    configure_logging()
    host = os.getenv("FLASK_HOST", "0.0.0.0")
    port = int(os.getenv("FLASK_PORT", "5000"))
    LOG.info("Starting Flask receiver on %s:%s", host, port)
    app.run(host=host, port=port)
